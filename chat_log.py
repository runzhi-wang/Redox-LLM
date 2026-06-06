"""Persist RAG Q&A sessions with user comments; admin XLSX export."""

from __future__ import annotations

import json
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import CHAT_LOG_PATH, OUTPUT_DIR

EXPORT_COLUMNS = [
    ("id", "record_id"),
    ("created_at", "timestamp_utc"),
    ("question", "question"),
    ("answer", "answer"),
    ("model", "model"),
    ("top_k", "top_k"),
    ("rating", "quality_rating"),
    ("comment", "user_comment"),
    ("comment_updated_at", "comment_updated_utc"),
    ("refs_summary", "retrieved_sources"),
]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_log_file() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not CHAT_LOG_PATH.exists():
        CHAT_LOG_PATH.write_text("", encoding="utf-8")


def _read_all() -> list[dict[str, Any]]:
    _ensure_log_file()
    records: list[dict[str, Any]] = []
    for line in CHAT_LOG_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        records.append(json.loads(line))
    return records


def _write_all(records: list[dict[str, Any]]) -> None:
    _ensure_log_file()
    CHAT_LOG_PATH.write_text(
        "\n".join(json.dumps(r, ensure_ascii=False) for r in records) + ("\n" if records else ""),
        encoding="utf-8",
    )


def _refs_summary(refs: list[dict]) -> str:
    parts = []
    for ref in refs:
        cite = ref.get("cite", "")
        dist = ref.get("distance")
        if dist is not None:
            parts.append(f"{cite} (dist={float(dist):.4f})")
        else:
            parts.append(cite)
    return " | ".join(parts)


def append_record(
    *,
    question: str,
    answer: str,
    model: str,
    top_k: int,
    refs: list[dict],
) -> dict[str, Any]:
    record = {
        "id": str(uuid.uuid4()),
        "created_at": _now_iso(),
        "question": question,
        "answer": answer,
        "model": model,
        "top_k": top_k,
        "rating": "",
        "comment": "",
        "comment_updated_at": "",
        "refs_summary": _refs_summary(refs),
        "refs": refs,
    }
    with CHAT_LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return record


def list_records(limit: Optional[int] = None) -> list[dict[str, Any]]:
    records = _read_all()
    records.sort(key=lambda r: r.get("created_at", ""), reverse=True)
    if limit is not None:
        return records[:limit]
    return records


def update_comment(record_id: str, comment: str) -> bool:
    return update_feedback(record_id, comment=comment)


def update_feedback(
    record_id: str,
    *,
    comment: Optional[str] = None,
    rating: Optional[str] = None,
) -> bool:
    records = _read_all()
    updated = False
    for record in records:
        if record.get("id") == record_id:
            if comment is not None:
                record["comment"] = comment.strip()
            if rating is not None:
                record["rating"] = rating
            record["comment_updated_at"] = _now_iso()
            updated = True
            break
    if updated:
        _write_all(records)
    return updated


def export_xlsx(output_path: Path) -> int:
    records = list_records()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "chat_logs"

    header_font = Font(bold=True)
    for col_idx, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, record in enumerate(records, start=2):
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if key in ("question", "answer", "comment", "refs_summary"):
                cell.alignment = wrap

    widths = {
        "id": 38,
        "created_at": 24,
        "question": 48,
        "answer": 64,
        "model": 14,
        "top_k": 8,
        "rating": 14,
        "comment": 36,
        "comment_updated_at": 24,
        "refs_summary": 56,
    }
    for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 18)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return len(records)
