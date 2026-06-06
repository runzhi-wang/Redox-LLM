"""Export all Chroma chunks to Excel (.xlsx), optionally with embedding vectors."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from chroma_utils import fetch_all_chunks
from config import CHROMA_DIR

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
DEFAULT_OUTPUT = OUTPUT_DIR / "chunks_export.xlsx"
MANIFEST_PATH = OUTPUT_DIR / "chunks_export_manifest.json"

# English headers only — avoids Excel mojibake on some Windows setups.
COLUMNS = [
    ("chunk_id", "chunk_id"),
    ("doi", "doi"),
    ("source_file", "source_file"),
    ("heading_path", "heading_path"),
    ("section", "section"),
    ("has_metrics", "has_metrics"),
    ("content_type", "content_type"),
    ("index_version", "index_version"),
    ("token_count", "token_count"),
    ("char_count", "char_count"),
    ("embedding_dim", "embedding_dim"),
    ("embedding", "embedding"),
    ("text", "text"),
]


def _format_embedding(vec: list[float] | None) -> str:
    if not vec:
        return ""
    rounded = [round(float(v), 6) for v in vec]
    return json.dumps(rounded, ensure_ascii=True, separators=(",", ":"))


def write_xlsx(rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "chunks"

    header_font = Font(bold=True)
    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            if key == "embedding":
                value = _format_embedding(row.get("embedding"))
            cell = ws.cell(row=1 + row_idx - 1, column=col_idx, value=value)
            if key in ("text", "embedding"):
                cell.alignment = wrap

    widths = {
        "chunk_id": 36,
        "doi": 28,
        "source_file": 32,
        "heading_path": 48,
        "section": 36,
        "has_metrics": 12,
        "content_type": 12,
        "index_version": 10,
        "token_count": 10,
        "char_count": 10,
        "embedding_dim": 12,
        "embedding": 48,
        "text": 80,
    }
    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 20)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def write_manifest(rows: list[dict], output_path: Path) -> None:
    with_vectors = sum(1 for row in rows if row.get("embedding"))
    manifest = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "row_count": len(rows),
        "with_vectors": with_vectors,
        "output_xlsx": str(output_path.resolve()),
        "unique_doi": len({row.get("doi", "") for row in rows if row.get("doi")}),
    }
    MANIFEST_PATH.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(description="Export Chroma chunks to Excel")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--with-vectors",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Include BGE-M3 embedding vectors (default: on)",
    )
    args = parser.parse_args()

    if not CHROMA_DIR.exists():
        print(f"Index missing: {CHROMA_DIR}. Run build_index.py first.")
        return 1

    output_path = args.output
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    rows = fetch_all_chunks(include_embeddings=args.with_vectors)
    if not rows:
        print("No chunks to export.")
        return 1

    write_xlsx(rows, output_path)
    write_manifest(rows, output_path)

    with_vectors = sum(1 for row in rows if row.get("embedding"))
    print(f"Exported {len(rows)} chunks -> {output_path.resolve()}")
    if args.with_vectors:
        print(f"Rows with embedding vectors: {with_vectors}/{len(rows)}")
    print(f"Manifest: {MANIFEST_PATH.resolve()}")
    print(f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
